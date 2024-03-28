# -----------------------------------------------------------------------------
# Title: Circularity Score Dashboard
# Author: Iris Reitsma, Diederik Hemmes
# Date: November 1, 2023
# Description: This script creates a streamlit dashboard. This dashboard calculates 
#   a circularity score given input parameters corresponding to a circular economy 
#   company. A user can select which expert input to incorporate in the calculation,
#   and can choose whether to score risk drivers using multiple variables or one
#   override. 
# -----------------------------------------------------------------------------

import streamlit as st
import pandas as pd
import datetime
import pytz 
from PIL import Image
# import gspread
# from google.oauth2 import service_account
import matplotlib.pyplot as plt
from matplotlib import colormaps
import numpy as np
from fpdf import FPDF
import tempfile
import openpyxl
# from pathlib import path
# import streamlit_authenticator as stauth

# users = ['diederik', 'jeroen']
# ids = ['dh', 'jj']
# pws = ['01', '02']

# authenticator = stauth.Authenticate(users, ids, pws, 'riskscorecard', 'cookieref', cookie_expiry_days=1)

# name, authentication_status, username = authenticator.login("Login", "main")

# if authentication_status == False:
#     st.error('Username/password is incorrect')

# if authentication_status == None:
#     st.error('Please enter your username and password')

# if authentication_status == True:

@st.cache_data
def load_images(image_path: str) -> None:
    """
    Load and display images in a formatted layout.

    This function loads and displays images from provided file paths in a
    structured layout using Streamlit's column layout features.

    Args:
        image_path (str): File path to the image.

    Returns:
        None
    """
    logos = Image.open(image_path)
    st.image(logos, width=400)

    return


@st.cache_data
def load_expert_data(filepath_weights: str) -> pd.DataFrame:
    """
    Load expert data from an Excel file.

    This function reads an Excel file containing expert weights data and returns
    the data as a pandas DataFrame.

    Args:
        filepath_weights (str): File path to the Excel file containing expert weights data.

    Returns:
        pd.DataFrame: DataFrame containing the loaded expert weights data.
    """
    weights_df = pd.read_excel(filepath_weights)

    return weights_df


@st.cache_data
def load_drivers_data(filepath_risk: str) -> pd.DataFrame:
    """
    Load risk drivers data from an Excel file.

    This function reads an Excel file containing risk drivers data and returns
    the data as a pandas DataFrame. 

    Args:
        filepath_risk (str): File path to the Excel file containing risk drivers data.

    Returns:
        pd.DataFrame: DataFrame containing the loaded risk drivers data with missing
        values filled using forward fill.
    """
    drivers_df = pd.read_excel(filepath_risk)
    drivers_df = drivers_df.fillna(method='ffill', axis=0)

    return drivers_df


@st.cache_data
def load_peak_data(filepath_peaks: str) -> pd.DataFrame:
    """
    Load peak extraction data from an Excel file.

    This function reads an Excel file containing peak extraction data and returns
    the data as a pandas DataFrame.

    Args:
        filepath_weights (str): File path to the Excel file containing peak extraction data.

    Returns:
        pd.DataFrame: DataFrame containing the loaded peak extraction data.
    """
    peaks_df = pd.read_excel(filepath_peaks)
    peaks_df = peaks_df.fillna('no examples')

    return peaks_df

    
def riskDriver(riskdriver_groupby: pd.DataFrame, rd_calc: pd.DataFrame, riskdriver: str, number: int, data_dict: dict) \
    -> tuple[float, str, list[float], list[str], float, str, str, float, dict]:
    """
    Calculate risk driver scores and provide user interface for customization.

    This function calculates risk driver scores based on user inputs for various
    variables. It presents an interactive user interface for selecting options
    and overwriting scores for each risk driver.

    Args:
        riskdriver_groupby (pd.DataFrame): Grouped data containing risk driver information.
        riskdriver (str): Name of the risk driver.
        number (int): Sequential number of the risk driver.

    Returns:
        tuple[float, str, list[float], list[str], float, str, str]: A tuple containing:
            - The calculated normalized score for the risk driver.
            - The score in text.
            - A list of normalized scores for variables.
            - A list of selected text score per variable.
            - The selected override score.
            - The selected override option in text.
            - The motivation of choosing an override.
            - The lowest possible relative score for the risk driver. 
    """
    # Define column width
    col1, col2 = st.columns([10, 5])

    # Retrieve driver information from groupby object
    variables = riskdriver_groupby['Variable'].unique().tolist()
    overrides_init = riskdriver_groupby['Overrides'].unique().tolist()
    overrides = ['Override: '+ i for i in overrides_init]

    riskdrivers_calc = rd_calc['Risk Driver'].unique().tolist()
    variables_calc = rd_calc[rd_calc['Risk Driver']==riskdriver]['Variable'].unique().tolist()

    # Create drop down for riskdriver
    with col1:
        with st.expander('Risk Driver ' + str(number) + ': ' + riskdriver):
            inputs = [None]*len(variables)
            inputs_text = [None]*len(variables)

            # Loop over variables
            min_score_rel = 0
            max_score = 0
            scored_points = 0
            for i, v in enumerate(variables):
                if v in (variables_calc):
                    # Generate dropdown and help information
                    options = riskdriver_groupby[riskdriver_groupby['Variable']==v]['Answers'].tolist()
                    examples = riskdriver_groupby[riskdriver_groupby['Variable']==v]['Example'].tolist()

                    print(v)
                    # print(options)
                    # print(examples)

                    string = generate_help(options, examples)
                    print(string)
                    choice = st.selectbox(v, options, help=string)

                    # Determine risk score for variable (v)
                    # v_score = options.index(choice) + 1
                    v_score = options.index(choice) 
                    v_score_rel = v_score / (len(options)-1)
                    
                    # Store variable score and which option was chosen for variable
                    inputs[i] = v_score_rel
                    inputs_text[i] = choice
                    scored_points += v_score
                    max_score += len(options)
                    # max_score += 1
                    # min_score_rel += 1 / len(options)
                    min_score_rel += 0
                else:
                    options = riskdriver_groupby[riskdriver_groupby['Variable']==v]['Answers'].tolist()
                    choice = st.selectbox(v, ['not applicable for selected business model'], help='not applicable for selected business model')
                    # st.write('not applicable for selected business model')
                    choice = 'not applicable for selected business model'
                    v_score = 0
                    v_score_rel = 0
                    
                    # Store variable score and which option was chosen for variable
                    inputs[i] = v_score_rel
                    inputs_text[i] = choice
                    scored_points += 0
                    max_score += 0
                    # min_score_rel += 1 / len(options)
                    min_score_rel += 0

                data_dict[v[0:3] + '_text'] = choice
                data_dict[v[0:3] + '_score'] = v_score
                data_dict[v[0:3] + '_relscore'] = v_score_rel



            if riskdriver in (riskdrivers_calc):
                # Determine score for risk driver
                score = sum(inputs)
                score_rel = score / len(variables_calc)
                min_score_rel = min_score_rel / len(variables_calc)
                score_rel_text = overrides_init[round((scored_points/max_score)*(len(overrides_init)))]

                # Add risk driver score to top of override box
                overrides.insert(0, score_rel_text)

            else:
                score_rel_text = 'not applicable'
                score = 0
                score_rel = 0
                min_score_rel = 0


    # Create dropdown for overwriting risk driver score
    with col2:
        override = False
        if riskdriver in (riskdrivers_calc): 
            override_text = st.selectbox('Override', overrides, label_visibility='collapsed')
            # override_text = st.selectbox('Not applicable', ['Risk Driver is not applicable to the selected business model'], label_visibility='collapsed')

                    # override risk score if override is selected
            if override_text != score_rel_text:
                score = overrides.index(override_text) 
                score_rel_override = score / (len(overrides_init))
                score_rel = score_rel_override
                # score_rel = convert_range(score_rel_override, 1/(len(overrides_init)), 1, min_score_rel, 1) # Scale override to same range as based variables
                score_rel_text = override_text
                override = True
        else:
            st.write('not applicable for selected business model')
            override_text = 'not applicable for selected business model'

    # Provide a motivation for the override
    motivation = None
    if override:
        motivation = st.text_area(label='Override motivation', placeholder='Please motivate why you selected an override for this risk driver.', key=number)
        
        # Display error if no motivation is provided
        if not motivation:
            st.error('Please provide a motivation before you continue', icon="🚨")

    data_dict[str(number) + '_override'] = override
    data_dict[str(number) + '_override_motivation'] = motivation
    data_dict[str(number) + '_totalscore'] = score
    data_dict[str(number) + '_totalrelscore'] = score_rel
    data_dict[str(number) + '_text'] = score_rel_text

    return score_rel, score_rel_text, inputs, inputs_text, override, override_text, motivation, min_score_rel, data_dict


@st.cache_data
def generate_help(options: list[str], examples: list[str]) -> str:
    """
    Generate a help string with examples for each option.

    This function generates a formatted help string containing examples for each
    option provided. The options and their corresponding examples are combined
    into a user-friendly format.

    Args:
        options (list[str]): A list of options for a variable.
        examples (list[str]): A list of examples corresponding to each option.

    Returns: 
        str: A formatted help string containing option names and examples.
    """ 
    help_string = 'These are examples of the answers: \n\n' 

    # Provide an example for each option
    for n, t in enumerate(options):
        help_string += t
        help_string += ': '
        help_string += """ \n """
        help_string += examples[n]
        help_string += '\n\n' 
    
    return help_string 


@st.cache_data
def convert_range(value: float, min_original: float, max_original: float, \
                min_new: float, max_new:float) -> float: 
    """
    Convert a value from one range to another range.

    This function takes a value within a specified original range and scales
    it to a new range defined by the given minimum and maximum values.

    Args:
        value (float): The value to be converted.
        min_original (float): The minimum value of the original range.
        max_original (float): The maximum value of the original range.
        min_new (float): The minimum value of the new range.
        max_new (float): The maximum value of the new range.

    Returns:
        float: The scaled value within the new range.
    """
    # Calculate the range of the original values
    range_original = max_original - min_original

    # Scale and shift the value to the new range
    scaled_value = (value - min_original) / range_original * (max_new - min_new) + min_new

    return scaled_value


@st.cache_data
def prepare_weights(weights_df: pd.DataFrame, finance_weight: bool, risk_weight: bool, 
                    invest_weight: bool, busdev_weight: bool, ondernemers_weight:bool) -> pd.DataFrame:
    """
    Prepare weighted scores based on selected expert groups.

    This function prepares the weights for different riskdrivers based on 
    user-selected expert groups. It calculates the weighted average of the selected 
    expert groups and returns the resulting weights for further analysis.

    Args:
        weights_df (pd.DataFrame): DataFrame containing expert groups and their scores.
        finance_weight (bool): If True, include Finance expert group in calculations.
        risk_weight (bool): If True, include Risk expert group in calculations.
        invest_weight (bool): If True, include Invest expert group in calculations.
        busdev_weight (bool): If True, include Business Development expert group in calculations.

    Returns:
        pd.DataFrame: A DataFrame containing calculated weights for each risk driver.
    """
    # Retrieve which weights are selected
    selected_groups = []
    if finance_weight:
        selected_groups.append('Sub score Finance')
    if risk_weight:
        selected_groups.append('Sub score Risk')
    if invest_weight:
        selected_groups.append('Sub score Invest')
    if busdev_weight:
        selected_groups.append('Sub score Bus Dev')
    if ondernemers_weight:
        selected_groups.append('Sub score Bus Dev')

    # Calculate the weighted average of the selected weights
    selected_weights_df = weights_df[selected_groups]
    weights_df['Sum'] = selected_weights_df.sum(axis=1)
    total_points = weights_df['Sum'].sum()
    weights_df['Weight'] = weights_df['Sum'] / total_points

    # Display expert weights dataframe
    # st.dataframe(weights_df)

    return weights_df 


def create_weights_UI() -> tuple[bool, bool, bool, bool]:
    """
    Create a user interface for choosing expert weights.

    This function generates a user interface with checkboxes for selecting expert
    weights from different expert groups. It allows users to choose which expert
    groups to include in the calculations.

    Returns:
        tuple[bool, bool, bool, bool]: A tuple containing Boolean values indicating whether
        each of the expert weights (Finance, Risk, Invest, Bus Dev) is selected.

    """
    st.header('Choose expert weights')

    # Create checkboxes for each expert group
    col1, col2, col3, col4 = st.columns([1,1,1,1.3])
    with col1:
        finance_weight = st.checkbox('Finance', value=True)
    with col2:
        risk_weight = st.checkbox('Risk', value=False)
    with col3:
        invest_weight = st.checkbox('Invest', value=False)
    with col4:
        busdev_weight = st.checkbox('Business Development', value=False)

    # At least one of the weights should be selected, display error if not
    if not any([finance_weight, risk_weight, invest_weight, busdev_weight]):
        st.error('Please select at least one expert weight', icon="🚨")

    return finance_weight, risk_weight, invest_weight, busdev_weight


def apply_weights(weights_df: pd.DataFrame, risk_driver: str, value: float) -> float:
    """
    Apply a weight to a given value based on a specific risk driver.

    This function applies a weight from a DataFrame of weights to a provided risk 
    driver score. It calculates the weighted value using the weight associated 
    with the given risk driver.

    Args:
        weights_df (pd.DataFrame): DataFrame containing risk driver weights.
        risk_driver (str): Name of the risk driver for which the weight should be applied.
        value (float): The value to be weighted.

    Returns:
        float: The value after applying the specified weight.
    """
    # Obtain corresponding weight
    weight = weights_df.loc[weights_df['Risk Factor'] == risk_driver, 'Weight_adj'].values[0]

    # Apply weight 
    weighted_value = value * weight

    return weighted_value


@st.cache_data
def plot_weights(weights_df_plot: pd.DataFrame) -> None:
    """
    Create a bar chart depicting the distribution of importance points over risk drivers,
    colored by expert group.

    Parameters:
        weights_df_plot (pandas.DataFrame): DataFrame containing importance weights per risk driver.

    Returns:
        None
    """
    # Normalize expert points
    column_sums = weights_df_plot.sum()
    normalized_df = weights_df_plot.iloc[:, 1:].div(column_sums, axis=1)
    normalized_df['Risk Factor'] = weights_df_plot['Risk Factor']

    # Set the 'Risk Factor' column as the index
    normalized_df.set_index('Risk Factor', inplace=True)

    # Get the column names and row indices
    columns = normalized_df.columns
    indices = normalized_df.index
    width = 0.1

    # Create an array for x-axis positions
    x = np.arange(len(indices))

    # Plotting the bar chart 
    plt.style.use('tableau-colorblind10')   
    fig, ax = plt.subplots(figsize=(10,5))
    for i, column in enumerate(columns):
        ax.bar(x + i * width, normalized_df[column], width=width, label=column) 

    # Adding labels, legend, and title 
    ax.set_xticks(x + (len(columns) - 1) * width / 2)
    ax.set_xticklabels(indices, rotation=45, ha='right', fontsize=12)
    ax.set_ylabel('Fraction of points', fontsize=14)
    ax.legend()

    # Show figure in dashboard
    st.pyplot(fig)

    return 


@st.cache_data
def stacked_bar_chart(data: list[float], labels: list[str]) -> None:
    """
    Generate a horizontal stacked bar chart for a single list of values using Matplotlib.

    Parameters:
        data (list[float]): A list of values to create the stacked bar chart.
        data (list[str]): A list of labels to use the stacked bar chart.

    Returns:
        None
    """

    # Plotting the bar chart 
    # plt.style.use('tableau-colorblind10')   
    plt.style.use('seaborn-v0_8-muted')   

    
    fig, ax = plt.subplots(figsize=(8, 2))
    left = np.zeros(len(labels))
    for i, value in enumerate(data):
        ax.barh([0], value, left=left, label=labels[i])
        left += value
        
    # Adding labels, legend, and title 
    ax.set_xlabel("Points", fontsize=14)
    ax.set_xlim(0, 100)
    ax.set_title("Build-up of score", fontsize=16)
    ax.set_xticklabels(ax.get_xticklabels(), fontsize=12)
    ax.axes.get_yaxis().set_visible(False)
    ax.legend(loc="center left", bbox_to_anchor=(1, 0.5), fontsize=12)

    # Show figure in dashboard
    st.pyplot(fig)


    return


#### main ##################################################################################################
if __name__ == '__main__':

    data_dict = {}
    # dict = {'key1':'geeks', 'key2':'for'} 
    # print("Current Dict is: ", dict) 
    
    # # using the subscript notation 
    # # Dictionary_Name[New_Key_Name] = New_Key_Value 
    
    # dict['key3'] = 'Geeks'
    # dict['key4'] = 'is'
    # dict['key5'] = 'portal'
    # dict['key6'] = 'Computer'

    version = '1.0'

    # Scorecard title
    st.title('Circular Risk Scorecard')

    st.subheader("version: " + version)

    # Create dashboard tabs
    tab1, tab2, tab3, tab4, tab5 = st.tabs(["Scorecard", "Read me", "Peak extraction years", "Distribution of expert weights", "Circular Business Models"])

    # # Set up the scope and credentials to Google Sheet
    # credentials = service_account.Credentials.from_service_account_info(
    #     st.secrets["gcp_service_account"],
    #     scopes=[
    #         "https://www.googleapis.com/auth/spreadsheets",
    #     ],
    # )

    # # Sheet link name in secrets.toml
    # sheet = 'sheet_url' 

    # # Try to open Google Sheet
    # try:
    #     client = gspread.authorize(credentials)
    #     sheets_url = st.secrets[sheet]
    #     sh = client.open_by_url(sheets_url)

    # # Display error if opening sheet fails
    # except:
    #     st.error('Please check your internet connection. If this is not the issue, the connection to the sheet is lost.', icon="🚨")

    # Obtain general information
    date = datetime.datetime.now(tz=pytz.timezone('Europe/Amsterdam')).date()
    time = datetime.datetime.now(tz=pytz.timezone('Europe/Amsterdam')).time().replace(microsecond=0)
    

    # Lists of banks 
    banks = ['ABN AMRO', 'Rabobank', 'ING', 'Other']

    # list of business models
    business_models = ['Product as a Service', 'Resource recovery (material sales model)', 'Circular supplies (product sales model)', 'Product life-time extension (service sales model)', 'Sharing platforms', 'Other']

    data_dict['Date'] = date
    data_dict['Time'] = time
    data_dict['version'] = version


#####################################################################################  
# Scorecard tab
#####################################################################################  
    with tab1:
        
        st.header('Fill in details')

        # # Create bank selection interface
        # chosen_bank = st.radio('Filled in by which bank', options=banks)
        # if chosen_bank == "Other":
        #     custom_option = st.text_input("Filled in by which bank:", placeholder='Bank name')
        #     chosen_bank = custom_option

        chosen_bank = ""

        # Create name and company input fields
        user = st.text_input(label='Filled in by', placeholder='Your financial institution name or code', key='name_key')
        client_company = st.text_input(label='Filled in for', placeholder='Name or code of company for which to determine a circular risk score', key='client_key')

        businessmodel_msg = "For more information on the business models see the \'Circular Business Models\' tab"

        # Create business model selection interface
        chosen_businessmodel = st.radio('Circular Business Models', options=business_models, help = businessmodel_msg)

        if chosen_businessmodel == "Other":
            motivation = st.text_area(label='Business model motivation', placeholder='Please motivate why you selected for a different business model.')
            # Display error if no motivation is provided
            if not motivation:
                st.error('Please provide a motivation before you continue', icon="🚨")

        # Store general information
        row = [str(date), str(time), version, chosen_businessmodel, user, client_company]
        text_row = [] # Row for storing variable selections in text format
        info_columns = len(row)

        data_dict['Institution'] = chosen_bank
        data_dict['User'] = user
        data_dict['Client_company'] = client_company
        data_dict['BusinessModel'] = chosen_businessmodel

        # Set up PDF for storing input
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size = 20)
        pdf.cell(200, 30, txt = 'Circular Risk Scorecard Summary', ln = 1, align = 'C')
        pdf.set_font("Arial", size = 14)
        pdf.cell(200, 10, txt = 'Filled in by: ' + user, ln = 1, align = 'L')
        pdf.cell(200, 5, txt = 'Filled in for: ' + client_company, ln = 1, align = 'L')

        # Read in expert weights file 
        filepath_weights = 'Data/Expert_weights.xlsx'
        weights_df_init = load_expert_data(filepath_weights)

        # Create expert weight selection UI
        # finance_weight, risk_weight, invest_weight, busdev_weight = create_weights_UI()
        finance_weight = risk_weight = invest_weight = busdev_weight = ondernemers_weight = True



        # Prepare weights based on selected expert groups
        weights_df = prepare_weights(weights_df_init, finance_weight, risk_weight, invest_weight, busdev_weight, ondernemers_weight)


        st.header('Score risk factors')

        # Read in risk drivers file
        filepath_risk = 'Data/Risk_drivers.xlsx'
        drivers_df = load_drivers_data(filepath_risk)

        # overrides_df = drivers_df

        drivers_df_calc = drivers_df[drivers_df[chosen_businessmodel] == 1]
        riskdrivers_calc = drivers_df_calc['Risk Driver'].unique().tolist()
        # print(riskdrivers_calc)

        weights_df['Weight_adj'] = 0
        # weights_df[weights_df['Risk Factor'].isin(riskdrivers_calc)]['Weight_adj'] = weights_df['Weight']/sum(weights_df[weights_df['Risk Factor'].isin(riskdrivers_calc)]['Weight'])
        weights_df.loc[weights_df['Risk Factor'].isin(riskdrivers_calc), 'Weight_adj'] = weights_df['Weight']/sum(weights_df[weights_df['Risk Factor'].isin(riskdrivers_calc)]['Weight'])

        # Create dropdown for each risk driver
        rd_number = 1
        risk_scores = []
        risk_drivers = []
        point_list = []

        # drivers_df = drivers_df[drivers_df['Risk Driver'].isin(['Ability of Management Team','Suitability for circular proposition'])]

        for name, rd in drivers_df.groupby('Risk Driver', sort=False):
            risk_drivers.append(name)

            rd_calc = drivers_df_calc
            
            # Create dropdown and override option
            risk, risk_text, inputs, inputs_text, override, override_text, motivation, min_score_rel, data_dict = riskDriver(rd, rd_calc, name, rd_number, data_dict)

            # Apply expert weights
            weighted_risk = apply_weights(weights_df, name, risk)

            # Store risk driver score scale for plotting buildup of final score
            weighted_min = apply_weights(weights_df, name, min_score_rel)
            weighted_max = apply_weights(weights_df, name, 1)
            points = convert_range(weighted_risk, min_original=weighted_min, max_original=weighted_max, min_new=weighted_max*100, max_new=0)
            point_list.append(points)

            # Store risk driver information in row for sheet
            risk_scores.append(weighted_risk)
            row.extend(inputs)
            row.append(risk)
            row.append(override)
            text_row.extend(inputs_text)
            text_row.append(risk_text)
            text_row.append(override_text)
            text_row.append(motivation)

            # Store risk driver information in PDF
            variables = rd['Variable'].unique().tolist()
            pdf.cell(200, 15, txt = "Risk driver " + str(rd_number) + ": " + name, ln = 1, align = 'L', )
            pdf.set_font("Arial", size = 12)
            if override: 
                pdf.cell(200, 10, txt = override_text, ln = 1, align = 'L')
                pdf.cell(200, 10, txt = 'override motivation: ' + motivation, ln = 1, align = 'L')
            else:
                for n, i in enumerate(inputs_text):
                    pdf.multi_cell(200, 5, txt = variables[n]  + ':', align = 'L')
                    pdf.set_font("Arial", size = 10)
                    pdf.cell(200, 10, txt = '   ' + i, ln = 1, align = 'L')
                    pdf.set_font("Arial", size = 12)
                pdf.cell(200, 10, txt='Risk Driver score: ' + risk_text , ln = 1, align = 'L')
            pdf.set_font("Arial", size = 14)
                
            rd_number += 1
            
            # else:
            #     risk, risk_text, inputs, inputs_text, override, override_text, motivation, min_score_rel = 'NA'

            #     row.extend(inputs)
            #     row.append(risk)
            #     row.append(override)
            #     text_row.extend(inputs_text)
            #     text_row.append(risk_text)
            #     text_row.append(override_text)
            #     text_row.append(motivation)
            


        st.header('Final Circular Risk Score')
        
        # Determine final score and convert in on scale 0 to 100
        final_score = sum(risk_scores)
        # min_score = 0.266246 # Lowest score possible when filling in form (select only expert group 'Finance')
        min_score = 0 # Lowest score possible when filling in form (select only expert group 'Finance')
        normalized_final_score = convert_range(final_score, min_original=min_score, max_original=1, min_new=100, max_new=0)

        # Display final circular risk score 
        col1, col2 = st.columns([3,2])
        with col1:
            st.metric('Circular Risk Score (low score = low risk): ', str(round(normalized_final_score)) + ' / 100')
            st.progress(round(normalized_final_score, 1)/100)

        # Request internal PD
        # with col2:
        #     internal_score = st.number_input('Internal PD (0.0100 = 1%)', value=1., min_value=0.000, max_value=1.000, step=0.0001, format="%.4f")
        internal_score = ""
        

        riskdrivers_calc = drivers_df_calc['Risk Driver'].unique().tolist()
        # print(risk_drivers)

        # index_riskdrivers = risk_drivers.index(riskdrivers_calc)

        indices = [i for i, e in enumerate(risk_drivers) if e in riskdrivers_calc]
        # print(indices)


        plot_riskdrivers = [risk_drivers[i] for i in indices]
        # print(plot_riskdrivers)
        plot_points = [point_list[i] for i in indices]
        # print(plot_points)

        # Show stacked bar chart of circular risk score build-up
        stacked_bar_chart(plot_points, plot_riskdrivers)

        # Store circular risk score and internal PD to row for sheet
        row.insert(info_columns, normalized_final_score)
        row.insert(info_columns+1, internal_score)
        row.extend([finance_weight, risk_weight, invest_weight, busdev_weight])



        # Store circular risk score to PDF
        pdf.set_font("Arial", size = 20)
        pdf.cell(200, 30, txt = 'Final Circular Risk Score: ' + str(round(normalized_final_score)) + ' / 100', ln = 1, align = 'C')

        st.header('Feedback and submit')

        # Request feedback and store in row for sheet
        feedback_score = st.text_area('Circular Risk Score', placeholder="Does the Circularity Risk Score match with your expectations of the risk of this company?")
        feedback_drivers = st.text_area('Risk Drivers', placeholder="Are there any risk drivers that you missed when filling in the scorecard?")
        feedback_UI = st.text_area('User-friendliness', placeholder="How easy was it to understand and fill in the scorecard?")
        feedback_open = st.text_area('Other feedback', placeholder='Please provide here any other feedback.')
        row.append(feedback_score)
        row.append(feedback_drivers)
        row.append(feedback_UI)
        row.append(feedback_open)

        # Add variable text input to row for sheet
        row.extend(text_row)

        # Create dashboard test box
        test = st.checkbox('Please select this box if you are testing the dashboard', value=False)
        row.insert(info_columns, test)

        # Create button to download data as PDF
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tf:
            pdf.output(tf.name)
        with open(tf.name, "rb") as f:
            st.download_button(label='Download scorecard as PDF', data=f.read(), file_name=client_company+' scorecard.pdf', mime='text/csv')

        # # Create button to submit data to google sheet
        # send = st.button("Submit")
        # if send:    
        #     st.success('Risk score submitted!')
        #     sh.sheet1.append_row(row)
            

        data_dict['finance_weight'] = finance_weight
        data_dict['risk_weight'] = risk_weight
        data_dict['invest_weight'] = invest_weight
        data_dict['busdev_weight'] = busdev_weight
        data_dict['ondernemers_weight'] = ondernemers_weight


        data_dict['Final_score'] = final_score
        data_dict['Final_score_normalized'] = normalized_final_score
        data_dict['feedback_score'] = feedback_score
        data_dict['feedback_drivers'] = feedback_drivers
        data_dict['feedback_UI'] = feedback_UI
        data_dict['feedback_open'] = feedback_open
        data_dict['test_run'] = test


        # initializing lists
        # colnames = ['col_' + str(i) for i in range(len(row))]
        colnames = [key for key in data_dict.keys()]

        colvalues = [val for val in data_dict.values()]
        
        # using dictionary comprehension
        # to convert lists to dictionary
        # mydict2 = [{colnames[i]: row[i] for i in range(len(colnames))}]
        # mydict2 = [data_dict]



        import csv
        
        # mydict2 = [{'Date':date, 'Time':time, 'Version':version, 'client_company':client_company, 'user':user}]
        
                # field names
        fields = colnames 

        
        # name of csv file
        filename = "university_records.csv"

            # writing to csv file
        with open(filename, 'w', newline='') as csvfile:
            # creating a csv dict writer object
            writer = csv.DictWriter(csvfile, fieldnames=fields, delimiter=';')
        
            # writing headers (field names)
            writer.writeheader()
        
            # writing data rows
            writer.writerows([data_dict])


        import email
        import email.mime.application
        import smtplib
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText
        import mimetypes

        import os
        from github import Github
        from io import StringIO

        # Define function to commit and push changes to GitHub
        def git_push(content, filename = "records_all.csv"):
            
            access_token = st.secrets["github_login"]["access_token"]
            g = Github(access_token)
            repo = g.get_repo(st.secrets["github_login"]["repo"])
            branch = st.secrets["github_login"]["branch"]

            # Fetch the current CSV file content
            file = repo.get_contents(filename, ref=branch)
            file_content = file.decoded_content.decode('utf-8')
            csv_file = StringIO(file_content)
            csv_reader = list(csv.reader(csv_file))

            # Append the new row to the CSV content
            csv_reader.append(content)

            # Convert the updated CSV content back to a string
            new_csv_content = StringIO()
            csv_writer = csv.writer(new_csv_content)
            csv_writer.writerows(csv_reader)
            new_csv_content = new_csv_content.getvalue()

            # Update the file on GitHub
            commit_message = "Added new data to the CSV file."
            repo.update_file(file.path, commit_message, new_csv_content, file.sha, branch=branch)
            print(f"Successfully updated '{filename}' with new data.")
            
        send_csv = st.button("Submit scorecard data to csv")
        if send_csv:
            try:
                git_push(data_dict.values(), "records_all.csv")
                st.success('Data submitted successfully to csv!')
            except:
                st.text('Try submitting again!')

            # st.success('Data submitted successfully!')


        send = st.button("Submit scorecard data")
        if send:    

            from_addr = "circularriskscorecard@outlook.com"
            to_addr = "circularriskscorecard@outlook.com"

            smtp_srv = "smtp-mail.outlook.com"

            msg = MIMEMultipart()
            msg['Subject'] = 'new data submission'
            msg['From'] = from_addr
            msg['To'] = to_addr

            # The main body is just another attachment
            body = MIMEText("""Data export from scorecard""")
            msg.attach(body)

            # PDF attachment
            filename='scorecard.csv'
            # fp=open(f,'rb')
            with open(csvfile.name, "rb") as f:
                fp = f
                att = email.mime.application.MIMEApplication(fp.read(),_subtype="csv")
                fp.close()
                att.add_header('Content-Disposition','attachment',filename=filename)
                msg.attach(att)

            smtp = smtplib.SMTP(smtp_srv,587)
            smtp.ehlo()
            smtp.starttls()
            smtp.ehlo()
            smtp.login(st.secrets.mail_login.user, st.secrets.mail_login.passwd)
            smtp.sendmail(from_addr, to_addr, msg.as_string())
            smtp.quit()

            st.success('Risk score submitted!')

#####################################################################################  
# Readme tab
#####################################################################################  
    with tab2:
        st.header('Background')
        st.markdown("""This dashboard outputs a circular risk score given input parameters\
                     corresponding to a circular economy company. This score can be used\
                     as a decision factor in deciding which circular economy deals are \
                    accepted or rejected. In addition, the dashboard can give insight into\
                     which risk factors are important in making these kinds of decisions. \
                    Please read the Circular Risk Score – a practical guide for more background information. """)

        st.header('Instructions')
        st.markdown('**Names**: It is possible to use dummy names for the \'filled \
                    in by\' and \'filled in for\' fields.')
        
        st.markdown("**Expert weights**: Different groups of experts distributed points over\
                    the different risk drivers. The weight distribution over the different \
                    risk drivers is visualized in the \'Distribution of expert weights\' tab.\
                    The score is built up with average weights from all expert groups."
        ) 
    
        st.markdown("""**Scoring risk drivers**: The scorecard contains six risk drivers, \
                    each with multiple variables. Place your mouse on the question mark \
                    icon for examples of how to rate a company. """
        ) 

        st.markdown("""**Override risk score**: If you are unsure about how to rate\
                    the variables, it is possible to provide an override for the\
                    risk driver score. This override replaces the score determined using \
                    the separate variables, which is displayed as the first option\
                    in the selection box""")

        st.markdown("**Circular risk score**: The form will show the determined \
                    circular risk score, which lies between 0 (low risk) and 100 (high risk).")

#####################################################################################  
# Peak Extraction Year tab
#####################################################################################          
        with tab3:    
            st.markdown("This table contains the expected peak extraction years for materials, which\
                        is needed for risk driver 3.1. Reference: \
                        [Sverdrup & Ragnarsdóttir (2014)](https://www.geochemicalperspectives.org/wp-content/uploads/v3n2.pdf).\
                        ")
            
            # Load and show table
            filepath_peaks = 'Data/Peak_extraction_years.xlsx'    
            peaks_df = load_peak_data(filepath_peaks)
            st.dataframe(peaks_df, hide_index=True, use_container_width=True)

#####################################################################################  
# Expert weight plot tab
#####################################################################################          
        with tab4:
            st.markdown("This bar chart shows what fraction of points was assigned to which risk driver \
                        for each expert group.")
            
            # Plot weights distribution
            plot_weights(weights_df_init[['Risk Factor', 'Sub score Risk', 'Sub score Bus Dev', 'Sub score Finance', 'Sub score Invest']])

#####################################################################################   
            
#####################################################################################  
# Business models tab
#####################################################################################  
            
        with tab5:
            st.header('Background on Circular Business Models')
            st.markdown("""The five circular business are derived from the Circular Economy Finance Guidelines (2018) - link: 
                        https://www.ing.com/Newsroom/News/ABN-AMRO-ING-and-Rabobank-launch-finance-guidelines-for-circular-economy.htm. """)

            st.header('Circular Business Models')

            st.markdown('**Resource recovery (material sales model)**: Businesses that sell circular materials (non-virgin and/or bio-based). \
                        An example is SusPhos that upcycles phosphate-rich waste streams to generate high-quality bio-based materials, \
                        as an alternative for fossil-based materials.')
            
            st.markdown('**Circular supplies (product sales model)**: Products made from either bio-based materials or non-virgin materials. \
                        An example is BE O lifestyle, that produces products such as reusable cups from bio-based materials (plants), \
                        which are fabricated by people with an occupational disability.')
            
            st.markdown('**Product life-time extension (service sales model)**: Services to extend the lifetime of a \
                        product (e.g. repairing, refurbishing). An example is a bicycle repair shop.')
            
            st.markdown('**Product as a Service**: Businesses that sell the output of the product according to the unit of use\
                         (e.g. per washing cycle) or the result (e.g. clean laundry). Examples are wash-as-a-service (Bundles), \
                        flowers-as-a-service (Reflower) and kitchen-as-a-service (Chainable).')
            
            st.markdown('**Sharing platforms**: A platform to share products between multiple users. \
                        An example is MyWheels, which shares electrical cars.')
            

        

            
    # Create logo header
    st.text(' ')
    st.text(' ')
    st.subheader('This dashboard was realized by:')
    image_path = 'Images/Logos4.png'
    load_images(image_path)




           

